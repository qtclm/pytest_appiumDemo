import openpyxl
from ruamel import  yaml
import os
import collections
from tool.Operation_logging import logs
# demjson:可以处理不规则的json字符串
# 常用的两个方法， 一个是 encode， 一个是 decode
# encode	将 Python 对象编码成 JSON 字符串
# decode	将已编码的 JSON 字符串解码为 Python 对象

class OperationYaml(object):
    # pyyaml
    # 应用最广泛
    # 封装的api不够简单
    # 不支持YAML1.2最新版
    # ruamel.yaml
    # 是pyyaml的衍生版
    # 封装的api简单
    # 支持YAML 1.2最新版

    def __init__(self, path='../config',file_path=None):
        yaml_path = os.path.join(os.path.dirname(__file__),path)
        if file_path == None:
            self.file_path = os.path.join(yaml_path, 'Commands.yaml')
        else:
            self.file_path = os.path.join(yaml_path, file_path)

        # self.data=self.read_data()

    # 获取所有数据
    def read_data(self, mode='r'):
        """ 读取yaml里面里面的数据"""
        try:
            with open(self.file_path, mode, encoding='utf8') as f:
                yaml_info = yaml.load(f, Loader=yaml.Loader)
                return yaml_info
        except Exception as error:
            print(f'读取yaml失败，错误如下：{error}')
            return False

    # @staictmethod:静态方法:
    # 将这个函数静态化了（不用写self参数了），也就是可以在不用实例化的前提下调用这个函数了,但是此时无法直接访问类实例与类方法
    # @classmethod:
    # 类方法,第一个参数为cls（类本身），可以直接访问类对象、方法，但是访问实例对象时必须先将类实例化

    # 通过key递归获取对应的值
    def readDataForKey(self, key=None, yaml_data=None):
        datas = self.read_data()
        # print(logs().out_varname(datas))
        # for循环字典这一层的所有key值
        if datas:
            if key:
                if yaml_data is None:
                    yaml_data = datas
                else:
                    yaml_data = yaml_data
                # print(yaml_data)
                for i in list(yaml_data.keys()):
                    # 如果当前的key是我们要找的
                    if i == key:
                        return yaml_data[i]
                    # 如果当前的key不是我们找的key，并且是字典类型
                    elif isinstance(yaml_data[i], dict):
                        # 使用递归方法，查找下一层的字典
                        return self.readDataForKey(key, yaml_data[i])
                    else:
                        continue
            else:
                pass
        else:
            return None

    # 只有一层的字典适用
    def readforKey_onetier(self, key=None):
        datas = self.read_data()
        if datas:
            if key:
                for i in datas.keys():
                    if i == key:
                        return datas[i]
                    else:
                        continue
            else:
                return None
        else:
            return None

    # 更新文件数据
    def write_yaml(self, data, mode='w'):
        """ 往yaml里面写入数据
            yamlFile：yaml文件名
            data：要写入的数据
            mode：写入方式： w，覆盖写入， a，追加写入
            将原数据读取出来，如果没有要加入的key，则创建一个，如果有，则执行key下面的数据修改
        """
        old_data = self.read_data()
        new_data = data
        if new_data == {} or new_data == [] or new_data == '':
            print('写入数据为空，跳过写入')
            return None
        if old_data and isinstance(old_data, dict):
            # collections.ChainMap - -将多个映射合并为单个映射
            # 如果有重复的键，那么会采用第一个映射中的所对应的值
            d = dict(collections.ChainMap(new_data, old_data))
            self.write_data(d)
            # 重新赋值给实例，不然数据写入的数据不会生效
        else:
            # 如果old_data为空，直接写入传入的数据
            self.write_data(new_data)

    # 写入数据基础方法
    def write_data(self, data, mode='w'):
        try:
            with open(self.file_path, mode, encoding="utf-8") as f:
                yaml.dump(data, f, Dumper=yaml.RoundTripDumper)
                return True
        except Exception as error:
            print(f'yaml文件写入失败，错误如下：\n{error}')
            return False


class OperationExcle(object):
    def __init__(self, file_address=None):
        self.dataCase_path = os.path.join(os.path.dirname(__file__),'../textCommon/run')
        if file_address is not None:
            self.file_address=os.path.join(self.dataCase_path,file_address)
            # 类实例化时即完成对工作薄对象进行实例化，访问写入的数据不能立马生效的问题
        else:
            self.file_address=os.path.join(self.dataCase_path,'DataCase_ALL.xlsx')
            # # 类实例化时即完成对工作薄对象进行实例化，访问写入的数据不能立马生效的问题
        if not os.path.exists(self.dataCase_path):
            os.makedirs(self.dataCase_path)
        if not os.path.exists(self.file_address):
            # 创建工作薄
            new_book = openpyxl.Workbook()
            new_book.save(self.file_address)
        self.workbook=openpyxl.load_workbook(self.file_address)
        
        # 将写入的数据存进集合，并集合存进list,方便多个类共享
        self.write_datas=set()
        self.write_list=[]
        self.data=self.get_data()

    #保存
    def save_workbook(self,file_address=None):
        if file_address is None:
            file_address=self.file_address
        else:
            try:
                file_address=os.path.join(self.dataCase_path,file_address)
            except BaseException as error:
                print("保存的文件路径还是不要乱来哦")
                raise Exception(error)
        self.workbook.save(file_address)
    
    # 将需要写入的数据写进实例，方便数据共享
    def writeDatasObject(self,value):
        self.write_datas.add(value)
        self.write_list=list(self.write_datas)
        
    # 获取工作簿对象并返回，property将方法转换为属性
    # 获取sheets的内容
    def get_data(self,sheet_id=None):
        # self.save_workbook()
        data=self.workbook
        if isinstance(sheet_id,int):
            sheet_id=sheet_id
        elif sheet_id is None:
            sheet_id=0
        else:
            raise TypeError("sheet_id类型错误")
        tables = data[(data.sheetnames[sheet_id])]
        # 返回数据前先保存关闭
        return tables

    # 获取单元格的行数
    def get_lines(self):
        tables = self.data
        '''max_row:获取行数，max_column:获取列数'''
        tables=tables.max_row
        return tables

    # 获取某一个单元格的内容
    def get_cell_value(self, row, col):
        cell_value=self.data.cell(row=row,column=col).value
        # 也可以使用：cell_value=self.data['A1'].value
        return cell_value
        # 根据行列返回表单内容

    # 写入数据
    def write_value(self, row, col, value):
        '''写入excle数据row，col，value'''
        # 设定单元格的值，三种方式
        # sheet.cell(row=2, column=5).value = 99
        # sheet.cell(row=3, column=5, value=100)
        # ws['A4'] = 4  # write
        data=self.get_data()
        data.cell(row=row,column=col).value=value
        self.save_workbook()

        
    
    # 根据对应的caseid找到对应行的内容
    def get_row_data(self, case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        return rows_data

    # 根据对应的caseid找到相应的行号
    def get_row_num(self, case_id):
        '''用例起始行为2，所以这里需要指定now初始值为2'''
        num = 2
        cols_data = self.get_cols_data(1)
        cols_data=[int(i) for i in cols_data]
        max_line=self.get_lines()
        try:
            case_id=int(case_id)
            if case_id<=max_line:
                if cols_data:
                    for col_data in cols_data:
                        if case_id == col_data:
                            return num
                        num = num + 1
                else:
                    return None
            else:
                print('依赖caseId不能大于用例总行数')
                return None
        except TypeError as typeerror:
            # print(typeerror)
            return None
            
    # 根据行号找到该行内容
    def get_row_values(self, row):
        cols=self.data.max_column#获取最大列
        rowdata=[]
        for i in range(1,cols+1):
            cellvalue=self.data.cell(row=row,column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 获取某一列的内容
    def get_cols_data(self, col=None):
        rows = self.data.max_row#获取最大行
        columndata=[]
        for i in range(2,rows+1):
            if col != None:
                cellvalue = self.data.cell(row=i,column=col).value
            else:
                cellvalue=self.data.cell(row=i,column=1).value
            columndata.append(cellvalue)
        return columndata

        


if __name__ == '__main__':
    dict1 = {'crm_course_name':{"is_refund_audit":333,'this':100}}
    # opym=OperationYaml()#'dependFieldInfo.yaml'  #'dependKeyInfo.yaml'
    # print(opym.readDataForKey('config'))
    opx=OperationExcle()
    # # print(opx.get_cols_data(7))
    # print(opx.get_row_num(11))
